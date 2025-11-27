"""Generate test Excel files with realistic tabular data starting at random rows.

This script creates 20 Excel spreadsheets with varied data themes and row offsets
to test the excel_to_parquet.py conversion tool's handling of real-world scenarios.
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font

fake = Faker()


def generate_header_row(theme: str) -> list[str]:
    """Return appropriate column headers for the data theme."""
    headers = {
        "financial_invoices": [
            "Invoice_ID",
            "Customer_Name",
            "Invoice_Date",
            "Due_Date",
            "Amount",
            "Tax",
            "Total",
            "Status",
            "Payment_Method",
        ],
        "financial_transactions": [
            "Transaction_ID",
            "Account_Number",
            "Transaction_Date",
            "Description",
            "Debit",
            "Credit",
            "Balance",
            "Category",
        ],
        "financial_balances": [
            "Account_ID",
            "Account_Name",
            "Account_Type",
            "Opening_Balance",
            "Credits",
            "Debits",
            "Closing_Balance",
            "Last_Updated",
        ],
        "inventory_products": [
            "SKU",
            "Product_Name",
            "Category",
            "Quantity_On_Hand",
            "Unit_Price",
            "Supplier",
            "Reorder_Level",
            "Last_Restocked",
        ],
        "inventory_warehouse": [
            "Location_Code",
            "Aisle",
            "Bin",
            "Product_SKU",
            "Quantity",
            "Unit_Cost",
            "Total_Value",
            "Last_Counted",
        ],
        "personnel_employees": [
            "Employee_ID",
            "First_Name",
            "Last_Name",
            "Department",
            "Position",
            "Hire_Date",
            "Salary",
            "Email",
            "Status",
        ],
        "personnel_departments": [
            "Department_ID",
            "Department_Name",
            "Manager",
            "Employee_Count",
            "Budget",
            "Cost_Center",
            "Location",
        ],
        "sales_orders": [
            "Order_ID",
            "Customer_Name",
            "Order_Date",
            "Product",
            "Quantity",
            "Unit_Price",
            "Total",
            "Shipping_Method",
            "Status",
        ],
        "sales_customers": [
            "Customer_ID",
            "Company_Name",
            "Contact_Person",
            "Email",
            "Phone",
            "Total_Orders",
            "Total_Revenue",
            "Last_Order_Date",
        ],
        "sales_revenue": [
            "Period",
            "Region",
            "Product_Line",
            "Units_Sold",
            "Revenue",
            "Cost_Of_Goods",
            "Gross_Profit",
            "Profit_Margin",
        ],
    }

    return headers.get(theme, headers["financial_invoices"])


def generate_data_rows(theme: str, num_rows: int, num_cols: int) -> list[list]:
    """Generate random but realistic data matching the theme."""
    rows = []

    for _ in range(num_rows):
        if theme == "financial_invoices":
            invoice_date = fake.date_between(start_date="-1y", end_date="today")
            due_date = invoice_date + timedelta(days=30)
            amount = round(random.uniform(100, 10000), 2)
            tax = round(amount * 0.08, 2)
            row = [
                f"INV-{fake.unique.random_int(10000, 99999)}",
                fake.company(),
                invoice_date.strftime("%Y-%m-%d"),
                due_date.strftime("%Y-%m-%d"),
                amount,
                tax,
                round(amount + tax, 2),
                random.choice(["Paid", "Pending", "Overdue", "Cancelled"]),
                random.choice(["Credit Card", "Bank Transfer", "Check", "Cash"]),
            ]

        elif theme == "financial_transactions":
            amount = round(random.uniform(10, 5000), 2)
            is_debit = random.choice([True, False])
            row = [
                f"TXN-{fake.unique.random_int(100000, 999999)}",
                fake.bban(),
                fake.date_between(start_date="-6m", end_date="today").strftime(
                    "%Y-%m-%d"
                ),
                fake.sentence(nb_words=4),
                amount if is_debit else 0,
                0 if is_debit else amount,
                round(random.uniform(1000, 50000), 2),
                random.choice(["Utilities", "Payroll", "Revenue", "Supplies", "Rent"]),
            ]

        elif theme == "financial_balances":
            opening = round(random.uniform(10000, 100000), 2)
            credits = round(random.uniform(1000, 20000), 2)
            debits = round(random.uniform(1000, 15000), 2)
            row = [
                f"ACC-{fake.random_int(1000, 9999)}",
                fake.bs(),
                random.choice(["Asset", "Liability", "Revenue", "Expense", "Equity"]),
                opening,
                credits,
                debits,
                round(opening + credits - debits, 2),
                fake.date_between(start_date="-30d", end_date="today").strftime(
                    "%Y-%m-%d"
                ),
            ]

        elif theme == "inventory_products":
            quantity = random.randint(0, 500)
            unit_price = round(random.uniform(5, 500), 2)
            row = [
                f"SKU-{fake.unique.random_int(10000, 99999)}",
                fake.catch_phrase(),
                random.choice(
                    ["Electronics", "Clothing", "Food", "Hardware", "Office"]
                ),
                quantity,
                unit_price,
                fake.company(),
                random.randint(10, 50),
                fake.date_between(start_date="-90d", end_date="today").strftime(
                    "%Y-%m-%d"
                ),
            ]

        elif theme == "inventory_warehouse":
            quantity = random.randint(1, 100)
            unit_cost = round(random.uniform(10, 200), 2)
            row = [
                f"{chr(65 + random.randint(0, 5))}{random.randint(1, 9)}",
                random.randint(1, 20),
                f"B{random.randint(1, 50)}",
                f"SKU-{fake.random_int(10000, 99999)}",
                quantity,
                unit_cost,
                round(quantity * unit_cost, 2),
                fake.date_between(start_date="-30d", end_date="today").strftime(
                    "%Y-%m-%d"
                ),
            ]

        elif theme == "personnel_employees":
            hire_date = fake.date_between(start_date="-10y", end_date="-1y")
            row = [
                f"EMP-{fake.unique.random_int(1000, 9999)}",
                fake.first_name(),
                fake.last_name(),
                random.choice(
                    ["Sales", "Engineering", "Marketing", "Finance", "HR", "Operations"]
                ),
                fake.job(),
                hire_date.strftime("%Y-%m-%d"),
                random.randint(40000, 150000),
                fake.company_email(),
                random.choice(["Active", "On Leave", "Terminated"]),
            ]

        elif theme == "personnel_departments":
            emp_count = random.randint(5, 50)
            row = [
                f"DEPT-{fake.random_int(100, 999)}",
                random.choice(
                    ["Sales", "Engineering", "Marketing", "Finance", "HR", "Operations"]
                ),
                fake.name(),
                emp_count,
                emp_count * random.randint(50000, 100000),
                f"CC-{fake.random_int(1000, 9999)}",
                fake.city(),
            ]

        elif theme == "sales_orders":
            order_date = fake.date_between(start_date="-6m", end_date="today")
            quantity = random.randint(1, 100)
            unit_price = round(random.uniform(10, 500), 2)
            row = [
                f"ORD-{fake.unique.random_int(10000, 99999)}",
                fake.company(),
                order_date.strftime("%Y-%m-%d"),
                fake.catch_phrase(),
                quantity,
                unit_price,
                round(quantity * unit_price, 2),
                random.choice(["Standard", "Express", "Overnight", "International"]),
                random.choice(["Pending", "Shipped", "Delivered", "Cancelled"]),
            ]

        elif theme == "sales_customers":
            total_orders = random.randint(1, 50)
            avg_order = round(random.uniform(500, 5000), 2)
            row = [
                f"CUST-{fake.unique.random_int(1000, 9999)}",
                fake.company(),
                fake.name(),
                fake.company_email(),
                fake.phone_number(),
                total_orders,
                round(total_orders * avg_order, 2),
                fake.date_between(start_date="-1y", end_date="today").strftime(
                    "%Y-%m-%d"
                ),
            ]

        elif theme == "sales_revenue":
            units = random.randint(100, 10000)
            revenue = round(units * random.uniform(20, 200), 2)
            cogs = round(revenue * random.uniform(0.4, 0.7), 2)
            gross_profit = round(revenue - cogs, 2)
            row = [
                f"Q{random.randint(1, 4)}-{random.randint(2022, 2024)}",
                random.choice(["North", "South", "East", "West", "Central"]),
                random.choice(
                    ["Electronics", "Clothing", "Food", "Hardware", "Office"]
                ),
                units,
                revenue,
                cogs,
                gross_profit,
                f"{round((gross_profit / revenue) * 100, 1)}%",
            ]

        else:
            # Default: generic data
            row = [fake.word() for _ in range(num_cols)]

        rows.append(row)

    return rows


def create_excel_file(output_path: Path, start_row: int, theme: str):
    """Create an Excel file with data starting at specified row.

    Args:
        output_path: Path to save the Excel file
        start_row: 1-indexed row where data table begins (1 = row 1, no offset)
        theme: Data theme identifier (e.g., 'financial_invoices')
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Generate headers and data
    headers = generate_header_row(theme)
    num_cols = len(headers)
    num_rows = random.randint(10, 50)
    data_rows = generate_data_rows(theme, num_rows, num_cols)

    # Add title/metadata rows before data if start_row > 1
    if start_row > 1:
        # Add some realistic pre-header content
        theme_display = theme.replace("_", " ").title()
        ws.cell(row=1, column=1, value=f"{theme_display} Report")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        if start_row > 2:
            ws.cell(
                row=2,
                column=1,
                value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            )

        if start_row > 3:
            ws.cell(row=3, column=1, value=f"Total Records: {num_rows}")

    # Write header row at start_row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Write data rows starting at start_row + 1
    for row_idx, data_row in enumerate(data_rows, start=start_row + 1):
        for col_idx, value in enumerate(data_row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save workbook
    wb.save(output_path)
    print(
        f"Created {output_path.name} (start_row={start_row}, theme={theme}, rows={num_rows})"
    )


def main():
    """Generate 20 test Excel files with varied configurations."""
    output_dir = Path("data/test_excel")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Define themes to cycle through
    themes = [
        "financial_invoices",
        "financial_transactions",
        "financial_balances",
        "inventory_products",
        "inventory_warehouse",
        "personnel_employees",
        "personnel_departments",
        "sales_orders",
        "sales_customers",
        "sales_revenue",
    ]

    # Generate 20 files with varied start rows
    print("Generating 20 test Excel files...\n")

    for i in range(20):
        # Cycle through themes
        theme = themes[i % len(themes)]

        # Random start row between 1 and 20
        start_row = random.randint(1, 20)

        # Create filename
        filename = f"test_{i + 1:02d}_{theme}.xlsx"
        output_path = output_dir / filename

        # Generate file
        create_excel_file(output_path, start_row, theme)

    print(f"\n✓ Successfully generated 20 Excel files in {output_dir}")
    print(f"\nTo test conversion, run:")
    print(f"  uv run python excel_to_parquet.py {output_dir} -o data/test_output")


if __name__ == "__main__":
    main()
